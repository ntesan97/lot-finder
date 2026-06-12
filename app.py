import streamlit as st
import re
import io
import openpyxl
from pdfminer.high_level import extract_text

st.set_page_config(page_title="Invoice Lot Number Enricher", page_icon="📋", layout="wide")

st.title("📋 Invoice Lot Number Enricher")
st.markdown("Upload a PDF invoice and an Excel file. The app matches items by description and appends lot/certificate info into the Excel Description column.")

TARGET_SHEET = "Edit - Posted Sales Invoice - "
DESC_COL = "Description"


def get_descriptions_from_excel(excel_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    if TARGET_SHEET not in wb.sheetnames:
        return None, f"Sheet '{TARGET_SHEET}' not found."
    ws = wb[TARGET_SHEET]
    header_row = None
    desc_col_idx = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == DESC_COL:
                header_row = cell.row
                desc_col_idx = cell.column
                break
        if header_row:
            break
    if not desc_col_idx:
        return None, f"Column '{DESC_COL}' not found."
    descs = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        val = row[desc_col_idx - 1]
        if val:
            descs.append(str(val).strip())
    return descs, None


def extract_lot_info(pdf_bytes, descriptions):
    pages = extract_text(io.BytesIO(pdf_bytes)).split("\x0c")
    # Some exports contain a duplicate copy of the invoice appended after
    # the totals; stop once we've captured one full copy of the document.
    kept_pages = []
    for page in pages:
        kept_pages.append(page)
        if "Ukupno RSD sa PDV-om" in page:
            break
    text = "\x0c".join(kept_pages)
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    br_re = re.compile(r"^Br\. serije:.+")

    lot_map = {}
    for desc in descriptions:
        try:
            pos = next(i for i, l in enumerate(lines) if l.strip() == desc)
        except StopIteration:
            continue
        lots = []
        # Collect br lines preceding the description (the PDF layout places
        # each item's "Br. serije:" lines directly before its description)
        j = pos - 1
        while j >= 0 and br_re.match(lines[j]):
            lots.insert(0, lines[j])
            j -= 1
        if lots:
            lot_map[desc] = " | ".join(lots)
    return lot_map


def enrich_excel(excel_bytes, lot_map):
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb[TARGET_SHEET]
    header_row = None
    desc_col_idx = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == DESC_COL:
                header_row = cell.row
                desc_col_idx = cell.column
                break
        if header_row:
            break

    matched, unmatched = [], []
    for row in ws.iter_rows(min_row=header_row + 1):
        cell = row[desc_col_idx - 1]
        desc = cell.value
        if not desc:
            continue
        desc_str = str(desc).strip()
        if desc_str in lot_map:
            cell.value = f"{desc_str}\n{lot_map[desc_str]}"
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
            matched.append(desc_str)
        else:
            unmatched.append(desc_str)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, matched, unmatched


col1, col2 = st.columns(2)
with col1:
    pdf_file = st.file_uploader("📄 Upload PDF Invoice", type=["pdf"])
with col2:
    excel_file = st.file_uploader("📊 Upload Excel File", type=["xlsx"])

if pdf_file and excel_file:
    pdf_bytes = pdf_file.read()
    excel_bytes = excel_file.read()

    with st.spinner("Reading Excel descriptions..."):
        descriptions, err = get_descriptions_from_excel(excel_bytes)
    if err:
        st.error(err)
        st.stop()

    with st.spinner("Extracting lot info from PDF..."):
        lot_map = extract_lot_info(pdf_bytes, descriptions)

    st.subheader("📦 Lot Info Extracted from PDF")
    if lot_map:
        for desc, lots in lot_map.items():
            st.markdown(f"**{desc}**")
            st.code(lots, language=None)
    else:
        st.warning("No lot info found in the PDF.")

    if lot_map:
        with st.spinner("Enriching Excel file..."):
            out_bytes, matched, unmatched = enrich_excel(excel_bytes, lot_map)

        st.subheader("✅ Results")
        col3, col4 = st.columns(2)
        with col3:
            st.success(f"**{len(matched)} items matched and updated**")
            for m in matched:
                st.markdown(f"- {m}")
        with col4:
            if unmatched:
                st.warning(f"**{len(unmatched)} items not matched in PDF**")
                for u in unmatched:
                    st.markdown(f"- {u}")
            else:
                st.success("All items matched!")

        st.download_button(
            label="⬇️ Download Enriched Excel",
            data=out_bytes,
            file_name=f"enriched_{excel_file.name}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
